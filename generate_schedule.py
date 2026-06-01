import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 列对应关系: col1=周一 ... col7=周日
weekday_map = {1: "周一", 2: "周二", 3: "周三", 4: "周四", 5: "周五", 6: "周六", 7: "周日"}

# 时间段
time_slots = [
    (1, "8:00", "8:50"),
    (2, "9:00", "9:50"),
    (3, "10:10", "11:00"),
    (4, "11:10", "12:00"),
    (5, "14:00", "14:50"),
    (6, "15:00", "15:50"),
    (7, "16:10", "17:00"),
    (8, "17:10", "18:00"),
    (9, "19:00", "19:50"),
    (10, "20:00", "20:50"),
    (11, "21:00", "21:50"),
    (12, "22:00", "22:50"),
]

# 课程数据: (星期列号, 起始节次, 结束节次, 课程名称, 上课地点)
courses = [
    # 周一
    (1, 1, 2, "计算机网络", "明德楼302"),
    (1, 3, 4, "静态网站设计", "统计分析与计算实验室(明德楼613)"),
    (1, 7, 8, "计算机网络", "数学建模实验室（明德楼611）"),
    # 周二
    (2, 1, 2, "静态网站设计", "软件项目开发实验室（明德楼405)"),
    (2, 7, 8, "马克思主义基本原理", "明理楼205"),
    # 周三
    (3, 1, 2, "马克思主义基本原理", "明理楼307"),
    (3, 3, 4, "离散数学", "耕读楼113"),
    # 周四
    (4, 3, 4, "马克思主义基本原理", "明理楼107"),
    # 周五
    (5, 3, 4, "马克思主义基本原理", "明仁楼304"),
    (5, 5, 8, "体育Ⅳ", "篮球场3"),
]

# 创建 Workbook
wb = openpyxl.Workbook()

# ===== 工作表1: 课程一览表 (按周次排列) =====
ws1 = wb.active
ws1.title = "课表视图"

# 样式定义
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

time_font = Font(bold=True, size=10)
time_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
time_align = Alignment(horizontal="center", vertical="center")

course_font = Font(size=10, bold=True)
loc_font = Font(size=9, color="666666")
course_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 颜色方案
color_fills = {
    "计算机网络": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "静态网站设计": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "马克思主义基本原理": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "离散数学": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "体育Ⅳ": PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
}

# 写标题行
ws1.cell(row=1, column=1, value="节次")
ws1.cell(row=1, column=2, value="时间")
for col in range(1, 8):
    cell = ws1.cell(row=1, column=col + 2, value=weekday_map[col])
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

ws1.cell(row=1, column=1).font = header_font
ws1.cell(row=1, column=1).fill = header_fill
ws1.cell(row=1, column=1).alignment = header_align
ws1.cell(row=1, column=1).border = thin_border
ws1.cell(row=1, column=2).font = header_font
ws1.cell(row=1, column=2).fill = header_fill
ws1.cell(row=1, column=2).alignment = header_align
ws1.cell(row=1, column=2).border = thin_border

# 构建一个查找表: (weekday_col, section_num) -> (course_name, location)
course_map = {}
for col, start, end, name, loc in courses:
    for s in range(start, end + 1):
        course_map[(col, s)] = (name, loc)

# 填充表格
for section_idx, (sec_num, start_time, end_time) in enumerate(time_slots):
    row = section_idx + 2
    
    # 节次编号
    c = ws1.cell(row=row, column=1, value=f"第{sec_num}节")
    c.font = time_font
    c.fill = time_fill
    c.alignment = time_align
    c.border = thin_border
    
    # 时间
    time_str = f"{start_time}-{end_time}"
    c = ws1.cell(row=row, column=2, value=time_str)
    c.font = time_font
    c.fill = time_fill
    c.alignment = time_align
    c.border = thin_border
    
    # 每一天的课程
    for weekday_col in range(1, 8):
        cell = ws1.cell(row=row, column=weekday_col + 2)
        cell.border = thin_border
        cell.alignment = course_align
        
        key = (weekday_col, sec_num)
        if key in course_map:
            name, loc = course_map[key]
            cell.value = f"{name}\n{loc}"
            cell.font = course_font
            if name in color_fills:
                cell.fill = color_fills[name]
        else:
            cell.value = ""
            cell.font = Font(size=10)

# 设置列宽
ws1.column_dimensions["A"].width = 10
ws1.column_dimensions["B"].width = 15
for col_letter in ["C", "D", "E", "F", "G", "H", "I"]:
    ws1.column_dimensions[col_letter].width = 28

# 行高
for r in range(2, len(time_slots) + 2):
    ws1.row_dimensions[r].height = 45

# ===== 工作表2: 课程明细表 =====
ws2 = wb.create_sheet("课程明细表")

# 表头
headers = ["星期", "上课时间", "课程名称", "上课地点"]
for i, h in enumerate(headers):
    cell = ws2.cell(row=1, column=i + 1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# 按星期排序
day_order = {"周一": 1, "周二": 2, "周三": 3, "周四": 4, "周五": 5, "周六": 6, "周日": 7}

# 生成明细数据
detail_data = []
for col, start, end, name, loc in courses:
    time_str = f"{time_slots[start-1][1]}-{time_slots[end-1][2]}"
    section_str = f"第{start}-{end}节" if start != end else f"第{start}节"
    time_desc = f"{weekday_map[col]} {section_str} ({time_str})"
    detail_data.append((weekday_map[col], time_desc, name, loc))

# 按星期排序
detail_data.sort(key=lambda x: day_order.get(x[0], 99))

for idx, (day, time_desc, name, loc) in enumerate(detail_data):
    row = idx + 2
    ws2.cell(row=row, column=1, value=day).border = thin_border
    ws2.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    ws2.cell(row=row, column=2, value=time_desc).border = thin_border
    ws2.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    
    ws2.cell(row=row, column=3, value=name).border = thin_border
    ws2.cell(row=row, column=3).font = Font(bold=True, size=11)
    
    ws2.cell(row=row, column=4, value=loc).border = thin_border
    
    # 交替行颜色
    if idx % 2 == 0:
        light_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
        for c in range(1, 5):
            ws2.cell(row=row, column=c).fill = light_fill

ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 22
ws2.column_dimensions["D"].width = 40

# ===== 工作表3: 课程汇总（按课程名） =====
ws3 = wb.create_sheet("课程汇总")

# 按课程名分组
from collections import defaultdict
grouped = defaultdict(list)
for col, start, end, name, loc in courses:
    time_str = f"{time_slots[start-1][1]}-{time_slots[end-1][2]}"
    grouped[name].append((weekday_map[col], time_str, loc))

headers3 = ["课程名称", "上课时间", "上课地点"]
for i, h in enumerate(headers3):
    cell = ws3.cell(row=1, column=i + 1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

row_idx = 2
for course_name in sorted(grouped.keys()):
    entries = grouped[course_name]
    start_row = row_idx
    for day, time_str, loc in entries:
        ws3.cell(row=row_idx, column=1, value=course_name).border = thin_border
        ws3.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
        ws3.cell(row=row_idx, column=2, value=f"{day} ({time_str})").border = thin_border
        ws3.cell(row=row_idx, column=3, value=loc).border = thin_border
        # 课程颜色
        if course_name in color_fills:
            for c in range(1, 4):
                ws3.cell(row=row_idx, column=c).fill = color_fills[course_name]
        row_idx += 1

ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 25
ws3.column_dimensions["C"].width = 45

# 保存
output_path = r"D:\xuexit\我的课表.xlsx"
wb.save(output_path)
print(f"✅ 课表已成功导出到: {output_path}")
print(f"\n📋 共导出 {len(detail_data)} 条课程记录")
print(f"\n📊 共 3 个工作表:")
print(f"   1. 课表视图 - 视觉课表布局")
print(f"   2. 课程明细表 - 逐条课程明细")
print(f"   3. 课程汇总 - 按课程名称分组")
