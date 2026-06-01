#!/usr/bin/env python3
"""
将静态HTML中保存的第14周课表导出为完整学期的Excel。
由于HTML只捕获了第14周的视图，而课程通常每周重复相同的时段，
本脚本导出两种格式：
  1. 周课表：每课按相同模式重复（可指定周范围）
  2. 明细表：包含周次信息的完整课程列表
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict

# ==================== 基础数据 ====================

# 学期信息
semester_info = "2025-2026 第2学期"
display_week = 14  # 第14周
total_weeks = 14   # 导出第1-14周

# 星期映射
weekday_map = {1: "周一", 2: "周二", 3: "周三", 4: "周四", 5: "周五", 6: "周六", 7: "周日"}

# 时间段
time_slots = [
    (1, "8:00", "8:50", "上午"),
    (2, "9:00", "9:50", "上午"),
    (3, "10:10", "11:00", "上午"),
    (4, "11:10", "12:00", "上午"),
    (5, "14:00", "14:50", "下午"),
    (6, "15:00", "15:50", "下午"),
    (7, "16:10", "17:00", "下午"),
    (8, "17:10", "18:00", "下午"),
    (9, "19:00", "19:50", "晚上"),
    (10, "20:00", "20:50", "晚上"),
    (11, "21:00", "21:50", "晚上"),
    (12, "22:00", "22:50", "晚上"),
]

# ==================== 课表数据（从第14周HTML提取） ====================
#
# 格式: (星期列号, 起始节次, 结束节次, 课程名称, 上课地点)
#
# 注意：这些课程在第14周出现在课表中，表示它们在运行。
# 每门课通常每周在同一时段重复，直到学期结束。
# 如果某门课只上部分周次（如单周/双周），请在下面手动设置 week_filter。

courses_base = [
    # 周一
    (1, 1, 2, "计算机网络", "明德楼302"),
    (1, 3, 4, "静态网站设计", "统计分析与计算实验室(明德楼613)"),
    (1, 7, 8, "计算机网络", "数学建模实验室（明德楼611）"),
    # 周二
    (2, 1, 2, "静态网站设计", "软件项目开发实验室（明德楼405)"),
    (2, 7, 8, "马克思主义基本原理", "明理楼205"),
    # 周三
    (3, 1, 2, "马克思主义基本原理", "明理楼307"),
    (3, 3, 4, "离散数学", "耕读楼113"),
    # 周四
    (4, 3, 4, "马克思主义基本原理", "明理楼107"),
    # 周五
    (5, 3, 4, "马克思主义基本原理", "明仁楼304"),
    (5, 5, 8, "体育Ⅳ", "篮球场3"),
]

# ==================== 周次配置 ====================
#
# 默认：所有课程每周都上（1-25周）
# 如果要设置特定课程的周次，请在此修改。
# 格式: (星期, 起始节次, 课程名称) -> "1-16" 或 "1,3,5,7,9,11,13,15" 等
#
# 典型配置：
# - 全周课程: weeks = "all" (1-25周)
# - 单周: weeks = "odd" (1,3,5,7,9,11,13,15,17,19,21,23,25)
# - 双周: weeks = "even" (2,4,6,8,10,12,14,16,18,20,22,24)
# - 指定周: weeks = "1-16" 或 weeks = "3,5,7,9"

week_config = {}

# 如果你知道某门课的具体周次，取消下面的注释并修改：
# week_config[(5, 5, "体育Ⅳ")] = "1-16"  # 体育课只上1-16周
# week_config[(1, 1, "计算机网络")] = "1-16"
# ... 等等

# ==================== 辅助函数 ====================

def get_weeks_for_course(col, section, name):
    """获取某门课的上课周次列表"""
    key = (col, section, name)
    if key in week_config:
        config = week_config[key]
        if config == "all":
            return list(range(1, total_weeks + 1))
        elif config == "odd":
            return [w for w in range(1, total_weeks + 1) if w % 2 == 1]
        elif config == "even":
            return [w for w in range(1, total_weeks + 1) if w % 2 == 0]
        elif "-" in config:
            parts = config.split("-")
            return list(range(int(parts[0]), int(parts[1]) + 1))
        else:
            return [int(x) for x in config.split(",")]
    else:
        # 默认：所有周
        return list(range(1, total_weeks + 1))


def get_week_range_str(weeks_list):
    """将周次列表转换为可读字符串，如 '1-16', '1-8,11-16'"""
    if not weeks_list:
        return ""
    sorted_weeks = sorted(set(weeks_list))
    ranges = []
    start = sorted_weeks[0]
    end = sorted_weeks[0]
    for w in sorted_weeks[1:]:
        if w == end + 1:
            end = w
        else:
            ranges.append(f"{start}-{end}" if start != end else str(start))
            start = w
            end = w
    ranges.append(f"{start}-{end}" if start != end else str(start))
    return "、".join(ranges)


# ==================== 生成完整课程列表（包含周次） ====================

all_courses = []  # (weekday_name, day_col, section_start, section_end, time_str, course_name, location, week_num)

for col, start, end, name, loc in courses_base:
    weeks_list = get_weeks_for_course(col, start, name)
    time_str = f"{time_slots[start-1][1]}-{time_slots[end-1][2]}"
    for week_num in weeks_list:
        all_courses.append((weekday_map[col], col, start, end, time_str, name, loc, week_num))

# ==================== 创建 Excel ====================

wb = openpyxl.Workbook()

# 样式
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

time_font = Font(bold=True, size=10)
time_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
time_align = Alignment(horizontal="center", vertical="center")

course_font = Font(size=10, bold=True)
loc_font = Font(size=9, color="666666")
course_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
week_font = Font(size=9, italic=True, color="888888")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 颜色方案
color_fills = {
    "计算机网络": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "静态网站设计": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "马克思主义基本原理": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "离散数学": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "体育Ⅳ": PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
}

# ==================== 工作表1: 课程明细表（含周次） ====================

ws1 = wb.active
ws1.title = "课程明细表(全学期)"

headers1 = ["周次", "星期", "节次", "上课时间", "课程名称", "上课地点"]
for i, h in enumerate(headers1):
    cell = ws1.cell(row=1, column=i + 1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

row_idx = 2
for entry in sorted(all_courses, key=lambda x: (x[7], x[1], x[2])):
    day_name, col, start, end, time_str, name, loc, week_num = entry
    section_str = f"第{start}-{end}节" if start != end else f"第{start}节"
    
    ws1.cell(row=row_idx, column=1, value=f"第{week_num}周").border = thin_border
    ws1.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    ws1.cell(row=row_idx, column=2, value=day_name).border = thin_border
    ws1.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center", vertical="center")
    
    ws1.cell(row=row_idx, column=3, value=section_str).border = thin_border
    ws1.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="center")
    
    ws1.cell(row=row_idx, column=4, value=time_str).border = thin_border
    ws1.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center", vertical="center")
    
    cell = ws1.cell(row=row_idx, column=5, value=name)
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border
    if name in color_fills:
        cell.fill = color_fills[name]
    
    ws1.cell(row=row_idx, column=6, value=loc).border = thin_border
    
    # 隔行配色
    if (week_num % 2) == 0:
        light_fill = PatternFill(start_color="F5F8FC", end_color="F5F8FC", fill_type="solid")
        for c in range(1, 7):
            if ws1.cell(row=row_idx, column=c).fill == PatternFill():  # 仅给无色的单元格
                ws1.cell(row=row_idx, column=c).fill = light_fill
    
    row_idx += 1

ws1.column_dimensions["A"].width = 12
ws1.column_dimensions["B"].width = 8
ws1.column_dimensions["C"].width = 14
ws1.column_dimensions["D"].width = 16
ws1.column_dimensions["E"].width = 22
ws1.column_dimensions["F"].width = 40

# 添加自动筛选
ws1.auto_filter.ref = f"A1:F{row_idx-1}"

# ==================== 工作表2: 每周课表视图 ====================
# 每页展示一周，逐周排列

ws2 = wb.create_sheet("每周课表视图")

current_row = 1
week_headers = ["节次", "时间", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

for week_num in range(1, total_weeks + 1):
    # 周标题
    title_cell = ws2.cell(row=current_row, column=1, value=f"=== 第{week_num}周 ({semester_info}) ===")
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    ws2.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    current_row += 1
    
    # 表头
    for i, h in enumerate(week_headers):
        cell = ws2.cell(row=current_row, column=i + 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    current_row += 1
    
    # 构建本周的课程查找表
    this_week_courses = {}
    for entry in all_courses:
        day_name, col, start, end, time_str, name, loc, w = entry
        if w == week_num:
            for s in range(start, end + 1):
                if s not in this_week_courses:
                    this_week_courses[s] = {}
                this_week_courses[s][col] = (name, loc)
    
    # 填充时间行
    for sec_num, start_time, end_time, period in time_slots:
        time_str = f"{start_time}-{end_time}"
        
        # 节次
        c = ws2.cell(row=current_row, column=1, value=f"第{sec_num}节")
        c.font = time_font
        c.fill = time_fill
        c.alignment = time_align
        c.border = thin_border
        
        # 时间
        c = ws2.cell(row=current_row, column=2, value=time_str)
        c.font = time_font
        c.fill = time_fill
        c.alignment = time_align
        c.border = thin_border
        
        # 每一天
        for weekday_col in range(1, 8):
            cell = ws2.cell(row=current_row, column=weekday_col + 2)
            cell.border = thin_border
            cell.alignment = course_align
            
            if sec_num in this_week_courses and weekday_col in this_week_courses[sec_num]:
                name, loc = this_week_courses[sec_num][weekday_col]
                cell.value = f"{name}\n{loc}"
                cell.font = course_font
                if name in color_fills:
                    cell.fill = color_fills[name]
            else:
                cell.value = ""
                cell.font = Font(size=10)
        
        current_row += 1
    
    # 留空行分隔
    current_row += 2

# 设置列宽
ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 15
for col_letter in ["C", "D", "E", "F", "G", "H", "I"]:
    ws2.column_dimensions[col_letter].width = 28

# ==================== 工作表3: 课程汇总（含周次信息） ====================

ws3 = wb.create_sheet("课程汇总")

# 按课程名分组
course_group = defaultdict(list)
for col, start, end, name, loc in courses_base:
    weeks_list = get_weeks_for_course(col, start, name)
    time_str = f"{time_slots[start-1][1]}-{time_slots[end-1][2]}"
    section_str = f"第{start}-{end}节" if start != end else f"第{start}节"
    course_group[name].append({
        "day": weekday_map[col],
        "time": time_str,
        "section": section_str,
        "loc": loc,
        "weeks": weeks_list,
    })

headers3 = ["课程名称", "上课时间", "节次", "上课地点", "上课周次"]
for i, h in enumerate(headers3):
    cell = ws3.cell(row=1, column=i + 1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

row_idx = 2
for course_name in sorted(course_group.keys()):
    entries = course_group[course_name]
    for entry in entries:
        ws3.cell(row=row_idx, column=1, value=course_name).border = thin_border
        ws3.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
        
        ws3.cell(row=row_idx, column=2, value=f"{entry['day']} ({entry['time']})").border = thin_border
        ws3.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
        
        ws3.cell(row=row_idx, column=3, value=entry['section']).border = thin_border
        ws3.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center")
        
        ws3.cell(row=row_idx, column=4, value=entry['loc']).border = thin_border
        
        week_str = get_week_range_str(entry['weeks'])
        ws3.cell(row=row_idx, column=5, value=f"第{week_str}周").border = thin_border
        
        if course_name in color_fills:
            for c in range(1, 6):
                ws3.cell(row=row_idx, column=c).fill = color_fills[course_name]
        
        row_idx += 1

ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 40
ws3.column_dimensions["E"].width = 20

# ==================== 保存 ====================

output_path = r"D:\xuexit\我的课表_全学期.xlsx"
wb.save(output_path)

# ==================== 统计输出 ====================

total_course_entries = len(all_courses)
unique_courses = len(course_group)
print(f"✅ 课表已成功导出到: {output_path}")
print(f"\n📊 统计信息:")
print(f"   - 学期: {semester_info}")
print(f"   - 总教学周数: {total_weeks} 周")
print(f"   - 课程数量: {unique_courses} 门")
print(f"   - 课表条目总数: {total_course_entries} 条")
print(f"\n📋 Excel 包含 {len(wb.sheetnames)} 个工作表:")
for s in wb.sheetnames:
    print(f"   📄 {s}")
print(f"\n⚙️ 周次配置说明:")
print(f"   所有课程默认按每周相同模式排课（第1-25周）。")
print(f"   如需为特定课程设置单周/双周或指定周次，")
print(f"   请修改脚本中的 week_config 字典后重新运行。")
